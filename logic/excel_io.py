from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


_HEADER_MAPPING = {
    "fecha": "fecha",
    "fecha programación": "fecha",
    "fecha programacion": "fecha",
    "referencia": "referencia",
    "color": "color",
    "cantidad": "cantidad",
}


def read_input_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    raw_headers = [cell.value for cell in ws[1]]
    headers = [str(h).lower().strip() if h else "" for h in raw_headers]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v for v in row if v is not None):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            key = _HEADER_MAPPING.get(h, h)
            row_dict[key] = row[i] if i < len(row) else None
        rows.append({
            "fecha": row_dict.get("fecha"),
            "referencia": str(row_dict.get("referencia", "") or "").strip(),
            "color": str(row_dict.get("color", "") or "").strip(),
            "cantidad": int(row_dict.get("cantidad", 0) or 0),
        })
    return rows


def _safe_fb_key(s):
    for ch in ['$', '#', '[', ']', '/', '.']:
        s = s.replace(ch, '-')
    return s.strip()


def read_referencias_excel(filepath):
    """Lee el template unificado de referencias (14 cols) y retorna dict para Firebase."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    referencias = {}
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        ref_a = str(row[0]).strip() if row[0] else ""
        color = str(row[3]).strip() if row[3] else ""
        if not ref_a or not color:
            continue
        fb_key = f"{_safe_fb_key(ref_a)}|{_safe_fb_key(color)}"
        referencias[fb_key] = {
            "referencia_a": ref_a,
            "referencia_b": str(row[1]).strip() if row[1] else "",
            "descripcion":  str(row[2]).strip() if row[2] else "",
            "color":        color,
            "color_num":    row[4],
            "medida":       str(row[5]).strip() if row[5] else "",
            "um":           str(row[6]).strip() if row[6] else "",
            "ref1":            str(row[7]).strip() if row[7] else "",
            "nombre_proceso1": str(row[8]).strip() if row[8] else "",
            "ref2_i":          str(row[9]).strip() if row[9] else "",
            "nombre_proceso2": str(row[10]).strip() if row[10] else "",
            "ref2_j":          str(row[11]).strip() if row[11] else "",
            "notas1":       str(row[12]).strip() if row[12] else "",
            "notas2":       str(row[13]).strip() if row[13] else "",
        }
    return referencias, errors


def create_referencias_template(target, data=None):
    """Crea el template unificado de referencias con 14 columnas. Si se pasa data, incluye los registros de Firebase."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Referencias"
    headers = [
        "Referencia", "Referencia PRD", "Descripción", "Color", "Color #",
        "Medida", "U.M",
        "Proceso 1", "Nombre Proceso 1",
        "Proceso 2", "Nombre Proceso 2",
        "REF2", "Notas Proceso 1", "Notas Proceso 2",
    ]
    _apply_headers(ws, headers, "1F4E79")
    if data:
        for r, (_key, d) in enumerate(data.items(), 2):
            ws.cell(row=r, column=1,  value=d.get("referencia_a", ""))
            ws.cell(row=r, column=2,  value=d.get("referencia_b", ""))
            ws.cell(row=r, column=3,  value=d.get("descripcion", ""))
            ws.cell(row=r, column=4,  value=d.get("color", ""))
            ws.cell(row=r, column=5,  value=d.get("color_num"))
            ws.cell(row=r, column=6,  value=d.get("medida", ""))
            ws.cell(row=r, column=7,  value=d.get("um", ""))
            ws.cell(row=r, column=8,  value=d.get("ref1", ""))
            ws.cell(row=r, column=9,  value=d.get("nombre_proceso1", ""))
            ws.cell(row=r, column=10, value=d.get("ref2_i", ""))
            ws.cell(row=r, column=11, value=d.get("nombre_proceso2", ""))
            ws.cell(row=r, column=12, value=d.get("ref2_j", ""))
            ws.cell(row=r, column=13, value=d.get("notas1", ""))
            ws.cell(row=r, column=14, value=d.get("notas2", ""))
    col_widths = [14, 14, 28, 14, 10, 10, 8, 12, 18, 12, 18, 10, 35, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(target)


def _apply_headers(ws, headers, fill_color):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 20


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def write_jumbo_excel(opp_list, target):
    from datetime import date
    today = date.today().strftime("%Y%m%d")
    wb = openpyxl.Workbook()

    # --- Hoja Documentos ---
    ws_doc = wb.active
    ws_doc.title = "Documentos"
    doc_headers = [
        "CONSECUTIVO DCTO", "FECHA AAAAMMDD", "PLANIFICADOR",
        "REF1", "REF2", "REF3", "NOTAS",
    ]
    _apply_headers(ws_doc, doc_headers, "1F4E79")
    for r, opp in enumerate(opp_list, 2):
        ws_doc.cell(row=r, column=1, value=opp["opp"])
        ws_doc.cell(row=r, column=2, value=today)
        ws_doc.cell(row=r, column=3, value=opp["planificador"])
        ws_doc.cell(row=r, column=4, value=opp["ref1"])
        ws_doc.cell(row=r, column=5, value=opp["ref2"])
        ws_doc.cell(row=r, column=6, value="")
        ws_doc.cell(row=r, column=7, value=opp["notas"])
    _auto_width(ws_doc)

    # --- Hoja Items ---
    ws_items = wb.create_sheet("Items")
    item_headers = [
        "NUMERO DCTO", "REGISTRO MVTO", "REFERENCIA", "EXT1", "EXT2",
        "U.M", "CANT PLANEADA", "FECHA INICIO AAAAMMDD", "FECHA TERMINACION AAAAMMDD",
        "METODO LISTA", "METODO RUTA", "MEDIDA REAL", "BODEGA",
    ]
    _apply_headers(ws_items, item_headers, "1F4E79")
    item_counter = {}
    for r, opp in enumerate(opp_list, 2):
        doc = opp["opp"]
        item_counter[doc] = item_counter.get(doc, 0) + 1
        ws_items.cell(row=r, column=1, value=doc)
        ws_items.cell(row=r, column=2, value=item_counter[doc])
        ws_items.cell(row=r, column=3, value=opp["referencia_item"])
        ws_items.cell(row=r, column=4, value=opp["ext1"])
        ws_items.cell(row=r, column=5, value=opp["ext2"])
        ws_items.cell(row=r, column=6, value=opp["um"])
        ws_items.cell(row=r, column=7, value=opp["cantidad"])
        ws_items.cell(row=r, column=8, value=opp["fecha_inicio"])
        ws_items.cell(row=r, column=9, value=opp["fecha_fin"])
        c_lista = ws_items.cell(row=r, column=10, value="0001")
        c_lista.number_format = "@"
        c_ruta = ws_items.cell(row=r, column=11, value="0001")
        c_ruta.number_format = "@"
        ws_items.cell(row=r, column=12, value="")
        ws_items.cell(row=r, column=13, value=opp.get("bodega", ""))
    _auto_width(ws_items)

    wb.save(target)


def create_input_template(target):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entrada"
    headers = ["Fecha Programación", "Referencia", "Color", "Cantidad"]
    _apply_headers(ws, headers, "2E75B6")
    ws.column_dimensions["A"].number_format = "DD/MM/YYYY"
    for i, w in enumerate([20, 15, 20, 12]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
    wb.save(target)


def create_estructura_template(target, data=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estructura"
    headers = ["Referencia", "Descripción", "Proceso 1", "Proceso 2", "Proceso 3",
               "Proceso 4", "Proceso 5", "Proceso 6", "Proceso 7", "Proceso 8"]
    _apply_headers(ws, headers, "2E75B6")
    if data:
        for r, (ref, info) in enumerate(data.items(), 2):
            procesos = info.get("procesos", [])
            row = [ref, info.get("descripcion", "")] + procesos
            for col, val in enumerate(row, 1):
                ws.cell(row=r, column=col, value=val)
    col_widths = [15, 28, 18, 18, 18, 18, 18, 18, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(target)


def read_estructura_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    referencias = {}
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        ref = str(row[0]).strip() if row[0] else ""
        if not ref or ref.lower() == "none":
            continue
        descripcion = str(row[1]).strip() if row[1] else ""
        procesos = [str(v).strip() for v in row[2:] if v and str(v).strip()]
        if not procesos:
            errors.append(f"Fila {row_idx}: '{ref}' no tiene procesos definidos — omitida.")
            continue
        referencias[ref] = {"descripcion": descripcion, "procesos": procesos}
    return referencias, errors
